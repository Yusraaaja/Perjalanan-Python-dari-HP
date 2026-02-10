import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

#1. Bikin datanya dulu seperti biasa
df = pd.read_excel('Laporan_Gaji_Januari.xlsx')
file_target = 'Laporan_Final.xlsx'
df.to_excel(file_target, index=False)

#2. Buka file nya dengan openpyxl untuk di hias
wb = load_workbook(file_target)
ws = wb.active

#3. Desain Header (Baris Pertama)
biru = PatternFill(start_color="00A86B", end_color="00A86B", fill_type="solid")
teks_putih = Font(color="FFFFFF", bold=True)
rata_tengah = Alignment(horizontal="center")

#Terapkan ke semua kolom dibaris 1
for cell in ws[1]:
    cell.fill = biru
    cell.font = teks_putih
    cell.alignment = rata_tengah

#4. Simpan hasil desain
wb.save(file_target)
print(f"Selesai! sekarang tampilan '{file_target}' keliatan bagus!")