import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Siapkan data mentah
data = {
    'No' : [1,2,3,4,5],
    'Tanggal' : ['01/02/2026', '02/02/2026', '03/02/2026', '04/02/2026', '05/02/2026'],
    'Keterangan' : ['Saldo Awal', 'Beli ATK', 'Bayar Listrik', 'Ambil Kas', 'Beli Printer'],
    'Debet' : [2000000, 0, 0, 1500000, 0],
    'Kredit' : [0, 350000, 150000, 0, 850000]
}

# 1. Bikin datanya dulu seperti biasa
df = pd.DataFrame(data)
file_target = 'Buku_Kas_Keuangan.xlsx'
df.to_excel(file_target, index=False, startrow=2) # Startrow=2 biar ada ruang buat judul di atas

# 2. Buka file nya dengan openpyxl untuk di hias
from openpyxl import load_workbook
wb = load_workbook(file_target)
ws = wb.active

# A. Membuat Judul (MERGE CELL)
ws.merge_cells('A1:E1')
ws['A1'] = "Laporan Keuangan Sederhana"
ws['A1'].font = Font(size=16, bold=True)
ws['A1'].alignment = Alignment(horizontal="center")

# B. Desain Header (Baris Pertama)
hijau_kas = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
putih_bold = Font(color="FFFFFF", bold=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

for cell in ws[3]: # Header ada dibaris 3 karna startrow=2 dipakai untuk judul
    cell.fill = hijau_kas
    cell.font = putih_bold
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# C. Memberikan border dan format rupiah ke semua data
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=5):
    for cell in row:
        cell.border = thin_border
        # Jika kolom Debet (4) atau Kredit (5), Kasih format angka titik
        if cell.column in [4, 5]:
            cell.number_format ='#,##0'

# 3. Simpan Hasilnya
wb.save(file_target)
print(f"Selesai! sekarang tampilan '{file_target}' keliatan bagus!")